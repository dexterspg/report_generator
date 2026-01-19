from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.worksheet import worksheet
import pandas as pd
import numpy as np
import time
from yaspin import yaspin
import os
from datetime import datetime

def nombre_formula(header: str, src_df : pd.DataFrame,  *args) -> pd.DataFrame:

    try:
        df=pd.DataFrame(columns=[header])
        df[header] = (
            src_df['Translation Type'] + " " +
            pd.to_datetime(
                src_df['Fiscal Year'].astype(str) + '-' + src_df['Fiscal Period'].astype(str)
            ).dt.strftime('%b-%Y').str.lower()
        )
        return df
    except Exception as e:
        print(e)

def descripcion_formula(header: str, src_df : pd.DataFrame,  *args) -> pd.DataFrame:
    df=pd.DataFrame(columns=[header])
    df[header] =  ('M1/' + src_df['Unit'].astype(str) + '/'  + src_df['Fiscal Year'].astype(str) + '/' + 
        src_df['Fiscal Period'].astype(str) + '/' + src_df['Contract Name'].astype(str) + '/' + 
        src_df['Vendor'].astype(str)
    )

    return df

def gl_account_split(header: str, src_df: pd.DataFrame, *args) -> pd.DataFrame:
    header_to_index = {
        "compania": 0,
        "Unidad": 1,
        "CC": 2,
        "ubicacion": 3,
        "cuenta": 4,
        "subcuenta": 5,
        "equipo": 6,
        "intercompania": 7,
    }
    
    idx = header_to_index.get(header)
    if idx is None:
        return pd.DataFrame()  # return an empty DataFrame if header not recognized
    
    splitted = src_df["GL Account"].str.split("-", expand=True)
    
    if splitted.shape[1] <= idx:
        col_data = pd.Series([None] * src_df.shape[0])
    else:
        col_data = splitted[idx]
    
    return pd.DataFrame({header: col_data})

formula_mappings = {
    "TC" : None, 
    "Nombre": nombre_formula,
    "divisa": lambda header, src_df, *args : pd.DataFrame({ header: src_df["Contract Currency"]}),
    "compania": gl_account_split,
    "Unidad": gl_account_split,
    "CC": gl_account_split,
    "ubicacion": gl_account_split,
    "cuenta": gl_account_split,
    "subcuenta": gl_account_split,
    "equipo": gl_account_split,
    "intercompania": gl_account_split,
    # "debito": lambda header, src_df ,*args : pd.DataFrame({header : src["Amount in Contract Currency"].astype(float) if }),
    "debito": lambda header, src_df, *args: pd.DataFrame({
    header: np.where(
        src_df["Amount in Contract Currency"].astype(float) > 0,
        src_df["Amount in Contract Currency"].astype(float),
        0
    )
}),
    "credito": lambda header, src_df, *args: pd.DataFrame({
    header: np.where(
        src_df["Amount in Contract Currency"].astype(float) < 0,
        np.abs(src_df["Amount in Contract Currency"].astype(float)),
        0
    )
}),
    "debito_convertido":"",
    "credito_convertido":"",
    "descripcion": descripcion_formula
}

cell_number_format = {
    "TC":'_-* #,##0.0000_-;-* #,##0.0000_-;_-* "-"??_-;_-@_-',
    "debito":'_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
    "credito": '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
    "debito_convertido": '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
    "credito_convertido":'_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
}


def _handle_formula(header: str, src_df,  *args) -> pd.DataFrame:
    formula_func = formula_mappings.get(header)
    
    if callable(formula_func):
        return formula_func(header, src_df, *args)  
    
    return pd.DataFrame()


def get_cell_formula_exch_multiplier(r_idx, tc_letter, debitcredit_letter):
    return f"=IF(OR(ISBLANK({tc_letter}{r_idx}),ISBLANK({debitcredit_letter}{r_idx})),0,{tc_letter}{r_idx}*{debitcredit_letter}{r_idx})"
    
def get_cell_formula_sum_rows(r_idx, debitcredit_letter):
    return f"=SUM({debitcredit_letter}2:{debitcredit_letter}{r_idx})"

def set_sum_cell(ws, last_row_idx, idx, letter, header ):
    cell = ws.cell(row = last_row_idx+1, column= idx+1)
    cell.value=get_cell_formula_sum_rows(last_row_idx, letter) 
    cell.number_format=cell_number_format[header]
    sum_fill= PatternFill(patternType="solid", fgColor="FFB6C1")  
    cell.fill= sum_fill

def apply_header_styling(ws, header_loc : int = 1):
    header_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(patternType="solid", fgColor="B3E5FC")  # Light blue color (adjust as needed)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_loc]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def mapped_data(source_file: str, updated_file: str, formula_mappings: dict,
                input_header_start: int = 27, input_data_start: int =28, template_header_start: int = 1,
                template_data_start: int = 2):
    start = time.perf_counter()
    print(f"⏳ Processing source file {source_file} ...")
    print("This may take a moment, please be patient ...")

    source_file = os.path.abspath(source_file)
        # template_file= os.path.abspath(template_file)
    updated_file= os.path.abspath(updated_file)

    input_df= pd.read_excel(source_file, header=input_header_start-1)
    # template_df = pd.read_excel(template_file) 

    # template_wb= load_workbook(template_file)
    # template_ws = template_wb.worksheets[0]

    template_df = pd.DataFrame(columns=[column for column in formula_mappings.keys()])
    for header in formula_mappings.keys():
        result_df :pd.DataFrame = _handle_formula(header, input_df, template_df)
        if not result_df.empty:
            template_df[header] = result_df[header]

    col_index_map = {
            header: template_df.columns.get_loc(header)
            for header in template_df.columns
        }

    tc_idx = int(col_index_map.get("TC"))
    tc_letter = get_column_letter(tc_idx + template_header_start)

    debito_idx = int(col_index_map.get("debito"))
    debito_letter = get_column_letter(debito_idx + 1)
    credito_idx = int(col_index_map.get("credito"))
    credito_letter = get_column_letter(credito_idx + 1)


    debito_convertido_idx = int(col_index_map.get("debito_convertido"))
    debito_converido_letter = get_column_letter(debito_convertido_idx + 1)
    credito_convertido_idx = int(col_index_map.get("credito_convertido"))
    credito_convertido_letter = get_column_letter(credito_convertido_idx + 1)

    template_wb = Workbook()
    template_ws = template_wb.active
    template_ws.title = "PolizaLedger"

    template_ws.append(template_df.columns.to_list())
    apply_header_styling(template_ws, template_header_start)


    print(f"Writing to excel file ... ")
    for r_idx, row in enumerate(template_df.itertuples(index=False), start=template_header_start):
        for header in template_df.columns:
            col_idx = col_index_map.get(header)
            if not isinstance(col_idx, int):
               raise ValueError(f"Header {header} not found in input data") 
            value = getattr(row, header)

            cell =template_ws.cell(
                row=r_idx+1,
                column=col_idx+1,
            )

            if header=="debito_convertido":
                cell.value=get_cell_formula_exch_multiplier(r_idx+1, tc_letter, debito_letter)

            elif header=="credito_convertido":
                cell.value=get_cell_formula_exch_multiplier(r_idx+1, tc_letter, credito_letter)

            else:
                cell.value = value

            if header in cell_number_format.keys():
                cell.number_format  = cell_number_format[header]


    last_row_idx = len(template_df) + template_header_start

    set_sum_cell(template_ws, last_row_idx, debito_idx, debito_letter, "debito")
    set_sum_cell(template_ws, last_row_idx, credito_idx, credito_letter, "credito")
    set_sum_cell(template_ws, last_row_idx, debito_convertido_idx, debito_converido_letter, "debito_convertido")
    set_sum_cell(template_ws, last_row_idx, credito_convertido_idx, credito_convertido_letter, "credito_convertido")

    # Save the updated workbook.
    template_wb.save(updated_file)
    end = time.perf_counter()
    print(f"Processing time completed: {end-start:.2f} seconds")


def main():

    # source_file = "Consolidated Transaction Report.xlsx"
    # template_file = "poliza_ledger_template.xlsx"

    source_file = input("Enter name of the excel file including extension  ex. file.xlsx\n(Recommended : Put the source file in the same folder as the script): ")
    current_time= datetime.now() 
    formatted_datetime = current_time.strftime("%Y%m%d_%H%M%S")
    # formatted_datetime = "" 
    updated_file = f"poliza_ledger_output{formatted_datetime}.xlsx"
    try:
        with yaspin(text="Loading... ", color="white") as spinner:
            time.sleep(2)  

            mapped_data(source_file, updated_file, formula_mappings, input_header_start=27,
                    input_data_start=28, template_header_start=1, template_data_start= 2)
            print(f"Output file saved to {updated_file}")
    except Exception as e:
        print(f"An error occurred: {e}")

    input("Press enter to exit .... ")

    
if __name__ == "__main__":
    main()



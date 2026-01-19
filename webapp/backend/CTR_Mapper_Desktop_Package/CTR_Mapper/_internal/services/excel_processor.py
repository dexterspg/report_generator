import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import os
from typing import Dict, Any
from .formula_mapper import FORMULA_MAPPINGS, CELL_NUMBER_FORMAT, handle_formula


def get_cell_formula_exch_multiplier(r_idx: int, tc_letter: str, debitcredit_letter: str) -> str:
    """Generate Excel formula for exchange rate multiplication"""
    return f"=IF(OR(ISBLANK({tc_letter}{r_idx}),ISBLANK({debitcredit_letter}{r_idx})),0,{tc_letter}{r_idx}*{debitcredit_letter}{r_idx})"


def get_cell_formula_sum_rows(r_idx: int, debitcredit_letter: str) -> str:
    """Generate Excel formula for summing rows"""
    return f"=SUM({debitcredit_letter}2:{debitcredit_letter}{r_idx})"


def set_sum_cell(ws, last_row_idx: int, idx: int, letter: str, header: str):
    """Set sum cell with formula and formatting"""
    cell = ws.cell(row=last_row_idx + 1, column=idx + 1)
    cell.value = get_cell_formula_sum_rows(last_row_idx, letter) 
    cell.number_format = CELL_NUMBER_FORMAT[header]
    sum_fill = PatternFill(patternType="solid", fgColor="FFB6C1")  
    cell.fill = sum_fill


def apply_header_styling(ws, header_loc: int = 1):
    """Apply styling to header row"""
    header_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(patternType="solid", fgColor="B3E5FC")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_loc]:  
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


class ExcelProcessor:
    """Main class for processing Excel files"""
    
    def __init__(self):
        self.formula_mappings = FORMULA_MAPPINGS
        
    def process_file(self, 
                     source_file_path: str, 
                     output_file_path: str,
                     input_header_start: int = 27, 
                     input_data_start: int = 28, 
                     template_header_start: int = 1,
                     template_data_start: int = 2) -> Dict[str, Any]:
        """
        Process the Excel file and return result information
        
        Args:
            source_file_path: Path to input Excel file
            output_file_path: Path for output Excel file
            input_header_start: Row number where headers start in input (1-indexed)
            input_data_start: Row number where data starts in input (1-indexed)
            template_header_start: Row number where headers start in output (1-indexed)
            template_data_start: Row number where data starts in output (1-indexed)
            
        Returns:
            Dictionary with processing results and metadata
        """
        try:
            # Read input Excel file
            input_df = pd.read_excel(source_file_path, header=input_header_start - 1)
            
            # Validate required columns
            required_columns = [
                'Translation Type', 'Fiscal Year', 'Fiscal Period', 'Unit',
                'Contract Name', 'Vendor', 'GL Account', 'Contract Currency',
                'Amount in Contract Currency'
            ]
            
            missing_columns = [col for col in required_columns if col not in input_df.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            # Create template DataFrame
            template_df = pd.DataFrame(columns=[column for column in self.formula_mappings.keys()])
            
            # Process each formula
            for header in self.formula_mappings.keys():
                result_df = handle_formula(header, input_df, template_df)
                if not result_df.empty:
                    template_df[header] = result_df[header]

            # Get column mappings
            col_index_map = {
                header: template_df.columns.get_loc(header)
                for header in template_df.columns
            }

            # Get specific column indices and letters
            tc_idx = int(col_index_map.get("TC"))
            tc_letter = get_column_letter(tc_idx + template_header_start)

            debito_idx = int(col_index_map.get("debito"))
            debito_letter = get_column_letter(debito_idx + 1)
            credito_idx = int(col_index_map.get("credito"))
            credito_letter = get_column_letter(credito_idx + 1)

            debito_convertido_idx = int(col_index_map.get("debito_convertido"))
            debito_convertido_letter = get_column_letter(debito_convertido_idx + 1)
            credito_convertido_idx = int(col_index_map.get("credito_convertido"))
            credito_convertido_letter = get_column_letter(credito_convertido_idx + 1)

            # Create workbook
            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws.title = "PolizaLedger"

            # Add headers
            template_ws.append(template_df.columns.to_list())
            apply_header_styling(template_ws, template_header_start)

            # Add data rows
            for r_idx, row in enumerate(template_df.itertuples(index=False), start=template_header_start):
                for header in template_df.columns:
                    col_idx = col_index_map.get(header)
                    if not isinstance(col_idx, int):
                        raise ValueError(f"Header {header} not found in input data") 
                    
                    value = getattr(row, header)
                    cell = template_ws.cell(row=r_idx + 1, column=col_idx + 1)

                    # Handle special formula columns
                    if header == "debito_convertido":
                        cell.value = get_cell_formula_exch_multiplier(r_idx + 1, tc_letter, debito_letter)
                    elif header == "credito_convertido":
                        cell.value = get_cell_formula_exch_multiplier(r_idx + 1, tc_letter, credito_letter)
                    else:
                        cell.value = value

                    # Apply number formatting
                    if header in CELL_NUMBER_FORMAT.keys():
                        cell.number_format = CELL_NUMBER_FORMAT[header]

            # Add sum totals
            last_row_idx = len(template_df) + template_header_start
            set_sum_cell(template_ws, last_row_idx, debito_idx, debito_letter, "debito")
            set_sum_cell(template_ws, last_row_idx, credito_idx, credito_letter, "credito")
            set_sum_cell(template_ws, last_row_idx, debito_convertido_idx, debito_convertido_letter, "debito_convertido")
            set_sum_cell(template_ws, last_row_idx, credito_convertido_idx, credito_convertido_letter, "credito_convertido")

            # Save the workbook
            template_wb.save(output_file_path)
            
            return {
                "success": True,
                "message": "File processed successfully",
                "input_rows": len(input_df),
                "output_rows": len(template_df),
                "output_file": output_file_path,
                "file_size": os.path.getsize(output_file_path)
            }
            
        except Exception as e:
            return {
                "success": False,
                "message": f"Processing error: {str(e)}",
                "error": str(e)
            }
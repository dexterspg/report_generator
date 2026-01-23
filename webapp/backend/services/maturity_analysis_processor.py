"""
Maturity Analysis Report Processor

Processes Consolidated Financial Schedules data and generates
maturity analysis reports with year buckets (Year 1-6 + Thereafter).

Based on logic from MaturityAnalysisReport.java
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from typing import Dict, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


# Output column headers (51 columns total)
OUTPUT_COLUMNS = [
    # Section 1: Contract/AG Metadata (1-24)
    'Erp System ID',
    'Contract ID',
    'Internal Contract Reference',
    'External Contract Reference',
    'Contract Name',
    'Company Code',
    'Business Unit',
    'Trading Partner ID',
    'Trading Partner Name',
    'Profit Center',
    'Cost Center',
    'Responsible Cost Center',
    'Activation Group ID',
    'Lease Classification',
    'Activation Group Status',
    'Accounting Start Date',
    'Likely Expiration Date',
    'Accounting Term In Months',
    'Accounting Term In Days',
    'Contract Currency',
    'Target Currency',
    'Exchange Rate To LC',
    'Exchange Rate To GC/RC',
    'Asset Class',
    # Section 2: Financial Values - Contract Currency (25-37)
    'Year 1',
    'Year 2',
    'Year 3',
    'Year 4',
    'Year 5',
    'Year 6',
    'Thereafter',
    'Total',
    'Less: Finance Charges',
    'Total Principal Liability',
    'ST Principal Closing Balance',
    'LT Principal Closing Balance',
    'Total Principal Liability',  # Duplicate name - ST + LT validation
    # Section 3: Financial Values - Target Currency (38-51)
    'Asset Class',  # Duplicate for target currency section
    'Year 1',
    'Year 2',
    'Year 3',
    'Year 4',
    'Year 5',
    'Year 6',
    'Thereafter',
    'Total',
    'Less: Finance Charges',
    'Total Principal Liability',
    'ST Principal Closing Balance',
    'LT Principal Closing Balance',
    'Total Principal Liability',  # Duplicate name - ST + LT validation
]

# Number format for financial columns
NUMBER_FORMAT = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
DATE_FORMAT = 'YYYY-MM-DD'


def calculate_period_bucket(payment_date: datetime, report_year: int, report_month: int) -> str:
    """
    Determine which year bucket a payment belongs to.

    The report date (year + month) is the reference point.
    Year 1 starts from the month AFTER the report date:
    - Year 1 = months 1-12 after report date
    - Year 2 = months 13-24 after report date
    - ... and so on
    - Thereafter = beyond 72 months (6 years)

    Args:
        payment_date: Date when payment is due (Period End Date)
        report_year: Report year (e.g., 2024)
        report_month: Report month (1-12, e.g., 12 for December)

    Returns:
        str: Bucket name ('skip', 'year1'-'year6', 'thereafter')
    """
    # Calculate months difference from report date to payment date
    months_diff = ((payment_date.year - report_year) * 12 +
                   (payment_date.month - report_month))

    if months_diff <= 0:
        return 'skip'      # Payment is at or before report date - ignore
    elif months_diff <= 12:
        return 'year1'     # Months 1-12 after report date
    elif months_diff <= 24:
        return 'year2'     # Months 13-24 after report date
    elif months_diff <= 36:
        return 'year3'     # Months 25-36
    elif months_diff <= 48:
        return 'year4'     # Months 37-48
    elif months_diff <= 60:
        return 'year5'     # Months 49-60
    elif months_diff <= 72:
        return 'year6'     # Months 61-72
    else:
        return 'thereafter' # Beyond 72 months


def calculate_term(start_date, end_date) -> Tuple[int, int]:
    """
    Calculate lease term in months and remaining days.

    Args:
        start_date: Lease start date
        end_date: Lease end date

    Returns:
        Tuple of (months, days)
    """
    if pd.isna(start_date) or pd.isna(end_date):
        return 0, 0

    try:
        # Convert to datetime if needed
        if not isinstance(start_date, datetime):
            start_date = pd.to_datetime(start_date)
        if not isinstance(end_date, datetime):
            end_date = pd.to_datetime(end_date)

        # Add 1 day to end_date for inclusive calculation
        rd = relativedelta(end_date + timedelta(days=1), start_date)
        months = rd.years * 12 + rd.months
        days = rd.days
        return months, days
    except Exception:
        return 0, 0


def safe_float(value, default=0.0) -> float:
    """Safely convert value to float."""
    if pd.isna(value):
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def parse_date(value) -> Optional[datetime]:
    """Safely parse date value."""
    if pd.isna(value):
        return None
    try:
        return pd.to_datetime(value)
    except Exception:
        return None


class MaturityAnalysisAggregator:
    """Aggregator for maturity analysis data by Contract + Activation Group."""

    def __init__(self):
        # Contract level aggregations (contract currency)
        self.year1 = defaultdict(float)
        self.year2 = defaultdict(float)
        self.year3 = defaultdict(float)
        self.year4 = defaultdict(float)
        self.year5 = defaultdict(float)
        self.year6 = defaultdict(float)
        self.thereafter = defaultdict(float)
        self.finance_charges = defaultdict(float)
        self.st_principal = defaultdict(float)
        self.lt_principal = defaultdict(float)

        # Contract metadata (first row wins)
        self.contract_info = {}

    def add_payment(self, key: str, bucket: str, amount: float):
        """Add payment amount to the appropriate bucket."""
        bucket_map = {
            'year1': self.year1,
            'year2': self.year2,
            'year3': self.year3,
            'year4': self.year4,
            'year5': self.year5,
            'year6': self.year6,
            'thereafter': self.thereafter
        }
        if bucket in bucket_map:
            bucket_map[bucket][key] += amount

    def add_finance_charge(self, key: str, amount: float):
        """Add finance charge (interest) amount."""
        self.finance_charges[key] += amount

    def set_principal_balances(self, key: str, st_principal: float, lt_principal: float):
        """Set principal balance values (only if not already set or if values are larger)."""
        # Take the values as they are extracted at period end
        self.st_principal[key] = st_principal
        self.lt_principal[key] = lt_principal

    def set_contract_info(self, key: str, info: Dict[str, Any]):
        """Set contract metadata (first occurrence only)."""
        if key not in self.contract_info:
            self.contract_info[key] = info

    def get_total(self, key: str) -> float:
        """Calculate total payments for a contract."""
        return (self.year1[key] + self.year2[key] + self.year3[key] +
                self.year4[key] + self.year5[key] + self.year6[key] +
                self.thereafter[key])

    def get_total_principal_from_payments(self, key: str) -> float:
        """Calculate total principal from payments minus finance charges."""
        return self.get_total(key) - self.finance_charges[key]

    def get_total_principal_from_balances(self, key: str) -> float:
        """Calculate total principal from ST + LT balances."""
        return self.st_principal[key] + self.lt_principal[key]


class MaturityAnalysisProcessor:
    """Main processor for maturity analysis reports."""

    def __init__(self):
        self.aggregator = None

    def process_file(
        self,
        source_file_path: str,
        output_file_path: str,
        report_year: int,
        report_month: int,
        exchange_rate: float = 1.0,
        input_header_start: int = 8,
        input_data_start: int = 9
    ) -> Dict[str, Any]:
        """
        Process the Excel file and generate maturity analysis report.

        The report date (year + month) is the reference point for year bucketing.
        Year 1 starts from the month after the report date.
        Target currency is taken from the "Company Currency" field in source data.

        Args:
            source_file_path: Path to Consolidated Financial Schedules.xlsx
            output_file_path: Path for output maturity analysis report
            report_year: Report year (e.g., 2024)
            report_month: Report month (1-12)
            exchange_rate: Exchange rate to apply (default 1.0)
            input_header_start: Row number where headers start (1-indexed)
            input_data_start: Row number where data starts (1-indexed)

        Returns:
            Dictionary with processing results
        """
        try:
            # Read input Excel file
            df = pd.read_excel(source_file_path, header=input_header_start - 1)

            # Validate required columns
            required_columns = [
                'Contract ID', 'Activation Group ID', 'Period End Date', 'Payment', 'Company Currency'
            ]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")

            # Process the data
            self.aggregator = MaturityAnalysisAggregator()
            self._process_rows(df, report_year, report_month, exchange_rate)

            # Generate output Excel (returns count of contracts included)
            contracts_included = self._generate_output(output_file_path, exchange_rate)

            return {
                "success": True,
                "message": "Maturity analysis report generated successfully",
                "input_rows": len(df),
                "output_rows": contracts_included,
                "total_contracts": len(self.aggregator.contract_info),
                "output_file": output_file_path,
                "file_size": os.path.getsize(output_file_path),
                "report_date": f"{report_year}-{report_month:02d}"
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"Processing error: {str(e)}",
                "error": str(e)
            }

    def _process_rows(
        self,
        df: pd.DataFrame,
        report_year: int,
        report_month: int,
        exchange_rate: float
    ):
        """Process all rows and aggregate data.

        Uses report year/month as reference point for year bucketing.
        Target currency is taken from Company Currency field.
        Only includes contracts/AGs that have payments after the report date.
        """

        for idx, row in df.iterrows():
            # Create aggregation key
            contract_id = str(row.get('Contract ID', ''))
            ag_id = str(row.get('Activation Group ID', ''))

            if not contract_id or not ag_id:
                continue

            key = f"{contract_id}_{ag_id}"

            # Parse dates
            start_date = parse_date(row.get('Activation Date'))
            end_date = parse_date(row.get('End Date'))

            # Store contract metadata (first occurrence)
            if key not in self.aggregator.contract_info:
                months, days = calculate_term(start_date, end_date)

                contract_currency = str(row.get('Contract Currency', ''))
                # Target currency comes from Company Currency field
                target_currency = str(row.get('Company Currency', ''))

                self.aggregator.set_contract_info(key, {
                    'erp_system_id': row.get('System', ''),
                    'contract_id': contract_id,
                    'internal_contract_reference': row.get('Internal Reference Number', ''),
                    'external_contract_reference': row.get('External Reference number', ''),
                    'contract_name': row.get('Contract Name', ''),
                    'company_code': row.get('Company Code', ''),
                    'business_unit': row.get('Business Unit', ''),
                    'trading_partner_id': row.get('Trading Partner', ''),
                    'trading_partner_name': '',  # Not in source
                    'profit_center': row.get('Unit Profit Center (Main)', ''),
                    'cost_center': row.get('Unit Cost Center (Main)', ''),
                    'responsible_cost_center': '',  # Not in source
                    'activation_group_id': ag_id,
                    'lease_classification': 'FINANCE',  # Default
                    'activation_group_status': row.get('Activation Group Status', ''),
                    'accounting_start_date': start_date,
                    'likely_expiration_date': end_date,
                    'accounting_term_months': months,
                    'accounting_term_days': days,
                    'contract_currency': contract_currency,
                    'target_currency': target_currency,
                    'exchange_rate_to_lc': 1.0,
                    'exchange_rate_to_gc': exchange_rate,
                    'asset_class': row.get('Unit Internal Asset Class', ''),
                })

            # Parse period end date for bucketing
            period_end_date = parse_date(row.get('Period End Date'))
            if not period_end_date:
                continue

            # Determine time bucket based on report date and add payment
            bucket = calculate_period_bucket(period_end_date, report_year, report_month)
            if bucket != 'skip':
                payment = safe_float(row.get('Payment', 0))
                self.aggregator.add_payment(key, bucket, payment)

            # Extract principal balances at the report date period
            if (period_end_date.year == report_year and
                period_end_date.month == report_month):
                st_principal = safe_float(row.get('ST Principal Liability Closing Balance', 0))
                lt_principal = safe_float(row.get('Principal Liability LT Closing Balance', 0))
                self.aggregator.set_principal_balances(key, st_principal, lt_principal)

            # Extract interest for periods after report date
            if (period_end_date.year > report_year or
                (period_end_date.year == report_year and
                 period_end_date.month > report_month)):
                interest = abs(safe_float(row.get('Interest Paid', 0)))
                self.aggregator.add_finance_charge(key, interest)

    def _generate_output(self, output_file_path: str, exchange_rate: float) -> int:
        """Generate the output Excel file.

        Returns:
            Number of contracts/AGs included in the output (those with payments > 0)
        """

        wb = Workbook()
        ws = wb.active
        ws.title = "Maturity Analysis"

        # Write headers
        headers = OUTPUT_COLUMNS
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(patternType="solid", fgColor="B3E5FC")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows (only include contracts with payments after report date)
        row_idx = 2
        contracts_included = 0
        for key, info in self.aggregator.contract_info.items():
            # Skip contracts/AGs that have no payments after report date
            total_cc = self.aggregator.get_total(key)
            if total_cc == 0:
                continue

            contracts_included += 1
            # Calculate values
            year1_cc = self.aggregator.year1[key]
            year2_cc = self.aggregator.year2[key]
            year3_cc = self.aggregator.year3[key]
            year4_cc = self.aggregator.year4[key]
            year5_cc = self.aggregator.year5[key]
            year6_cc = self.aggregator.year6[key]
            thereafter_cc = self.aggregator.thereafter[key]
            total_cc = self.aggregator.get_total(key)
            finance_charges_cc = -self.aggregator.finance_charges[key]  # Negative
            total_principal_cc = total_cc + finance_charges_cc
            st_principal_cc = self.aggregator.st_principal[key]
            lt_principal_cc = self.aggregator.lt_principal[key]
            total_principal_balance_cc = st_principal_cc + lt_principal_cc

            # Get exchange rate for this contract
            ex_rate = info.get('exchange_rate_to_gc', exchange_rate)

            # Target currency values
            year1_tc = year1_cc * ex_rate
            year2_tc = year2_cc * ex_rate
            year3_tc = year3_cc * ex_rate
            year4_tc = year4_cc * ex_rate
            year5_tc = year5_cc * ex_rate
            year6_tc = year6_cc * ex_rate
            thereafter_tc = thereafter_cc * ex_rate
            total_tc = total_cc * ex_rate
            finance_charges_tc = finance_charges_cc * ex_rate
            total_principal_tc = total_principal_cc * ex_rate
            st_principal_tc = st_principal_cc * ex_rate
            lt_principal_tc = lt_principal_cc * ex_rate
            total_principal_balance_tc = total_principal_balance_cc * ex_rate

            # Build row data (51 columns)
            row_data = [
                # Section 1: Metadata (1-24)
                info.get('erp_system_id', ''),
                info.get('contract_id', ''),
                info.get('internal_contract_reference', ''),
                info.get('external_contract_reference', ''),
                info.get('contract_name', ''),
                info.get('company_code', ''),
                info.get('business_unit', ''),
                info.get('trading_partner_id', ''),
                info.get('trading_partner_name', ''),
                info.get('profit_center', ''),
                info.get('cost_center', ''),
                info.get('responsible_cost_center', ''),
                info.get('activation_group_id', ''),
                info.get('lease_classification', ''),
                info.get('activation_group_status', ''),
                info.get('accounting_start_date'),
                info.get('likely_expiration_date'),
                info.get('accounting_term_months', 0),
                info.get('accounting_term_days', 0),
                info.get('contract_currency', ''),
                info.get('target_currency', ''),
                info.get('exchange_rate_to_lc', 1.0),
                info.get('exchange_rate_to_gc', 1.0),
                info.get('asset_class', ''),
                # Section 2: Contract Currency (25-37)
                year1_cc,
                year2_cc,
                year3_cc,
                year4_cc,
                year5_cc,
                year6_cc,
                thereafter_cc,
                total_cc,
                finance_charges_cc,
                total_principal_cc,
                st_principal_cc,
                lt_principal_cc,
                total_principal_balance_cc,
                # Section 3: Target Currency (38-51)
                info.get('asset_class', ''),  # Duplicate
                year1_tc,
                year2_tc,
                year3_tc,
                year4_tc,
                year5_tc,
                year6_tc,
                thereafter_tc,
                total_tc,
                finance_charges_tc,
                total_principal_tc,
                st_principal_tc,
                lt_principal_tc,
                total_principal_balance_tc,
            ]

            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply number format for financial columns (25-37 and 39-51)
                if 25 <= col_idx <= 37 or 39 <= col_idx <= 51:
                    cell.number_format = NUMBER_FORMAT

                # Apply date format for date columns (16-17)
                if col_idx in [16, 17] and value:
                    cell.number_format = 'YYYY-MM-DD'

            row_idx += 1

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max(12, len(str(header)) + 2)

        # Save workbook
        wb.save(output_file_path)

        return contracts_included
